"""
General Purpose
- Create an Excel load-order sheet from the `imports` array of a Vault loader config JSON.
- Preserve the original order of `imports` entries and map rows to a fixed business template.
- Apply Excel table formatting to improve filtering and readability.

Input Prerequisites
- JSON config file containing an `imports` array with fields `params`,
    `business_object_type`, `active`, and `status`.
- Python environment with `openpyxl` installed.

Output
- Excel file containing columns:
    `LoadOrderNr`, `ObjectType`, `Business ObjectType`, `ReferencedObject(s)`,
    `Load is Active`, `FunctionStatus`, `LoaderConfig`.
- If target file is locked, a timestamped fallback file is created automatically.

Start Parameter
- `--config` (optional): path to source Vault loader config JSON.
- `--output` (optional): path to destination Excel file.

Function
- `get_object_type_from_params`: extracts object type from the second token of `params`.
- `create_workbook_rows`: transforms imports entries into ordered output rows.
- `add_table`: applies Excel table style and filter-capable table range.
- `build_loadorder_excel`: orchestrates JSON loading, row creation, and file save.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


COLUMNS = [
    "LoadOrderNr",
    "LoaderFileName",
    "ObjectType",
    "Business ObjectType",
    "ReferencedObject(s)",
    "Load is Active",
    "FunctionStatus",
    "LoaderConfig",
    "BaseImportPath",
]


def get_object_type_from_params(params: Any) -> str:
    if not isinstance(params, str):
        return ""
    tokens = params.split()
    if len(tokens) < 2:
        return ""
    return tokens[1]


def get_loader_file_name_from_params(params: Any) -> str:
    if not isinstance(params, str):
        return ""
    match = re.search(r"(?:^|\s)-csv\s+(.+)$", params.strip())
    return match.group(1).strip() if match else ""


def load_json(file_path: Path) -> dict[str, Any]:
    with file_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def get_project_root() -> Path:
    script_dir = Path(__file__).resolve().parent
    for candidate in [script_dir, *script_dir.parents]:
        if (candidate / "config" / "vault_loader_config_basis_new.json").exists():
            return candidate
    raise FileNotFoundError("Could not locate project root containing config/vault_loader_config_basis_new.json")


def create_workbook_rows(
    imports: list[dict[str, Any]], loader_config_name: str, base_import_path: str
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for index, item in enumerate(imports, start=1):
        rows.append(
            [
                index,
                get_loader_file_name_from_params(item.get("params", "")),
                get_object_type_from_params(item.get("params", "")),
                item.get("business_object_type", ""),
                "",
                item.get("active", ""),
                item.get("status", ""),
                loader_config_name,
                base_import_path,
            ]
        )
    return rows


def add_table(ws, row_count: int, col_count: int) -> None:
    if row_count < 2:
        return
    from openpyxl.utils import get_column_letter

    end_col = get_column_letter(col_count)
    table_ref = f"A1:{end_col}{row_count}"
    table = Table(displayName="LoadorderTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_loadorder_excel(config_path: Path, output_path: Path) -> tuple[int, Path]:
    data = load_json(config_path)
    imports = data.get("imports", [])
    if not isinstance(imports, list):
        raise ValueError("Invalid config: 'imports' must be a list.")
    import_settings = data.get("import_settings", {})
    base_import_path = str(import_settings.get("import_path", "")).strip() if isinstance(import_settings, dict) else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Loadorder"

    ws.append(COLUMNS)

    rows = create_workbook_rows(imports, config_path.name, base_import_path)
    for row in rows:
        ws.append(row)

    add_table(ws, row_count=1 + len(rows), col_count=len(COLUMNS))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    saved_path = output_path
    try:
        wb.save(saved_path)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        saved_path = output_path.with_name(f"{output_path.stem}_locked_{timestamp}{output_path.suffix}")
        wb.save(saved_path)

    return len(rows), saved_path


def main() -> None:
    project_root = get_project_root()

    parser = argparse.ArgumentParser(
        description="Create load-order Excel from imports array in a vault loader config JSON."
    )
    parser.add_argument(
        "--config",
        default=str(project_root / "config" / "vault_loader_config_basis_new.json"),
        help="Path to loader config JSON file.",
    )
    parser.add_argument(
        "--output",
        default=str(project_root.parent / "master-data" / "Loadorder__2026-04.xlsx"),
        help="Path to output Excel file.",
    )
    args = parser.parse_args()

    config_path = Path(args.config).resolve()
    output_path = Path(args.output).resolve()

    row_count, saved_path = build_loadorder_excel(config_path, output_path)
    print(f"Created: {saved_path}")
    print(f"Rows written: {row_count}")


if __name__ == "__main__":
    main()
