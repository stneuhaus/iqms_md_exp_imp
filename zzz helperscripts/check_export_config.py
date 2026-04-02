"""
General Purpose
- Compare aggregated CSV attributes against export definitions in a Vault loader JSON config.
- Synchronize differences by appending missing CSV attributes to JSON export columns.
- Add JSON-only attributes into available HEAD columns in-memory for review.
- Generate a color-coded Excel report and an unmatched object-type CSV report.

Input Prerequisites
- A CSV file with columns: `object_type`, `file_intent`, and `HEAD1..HEADn`.
- A JSON config file containing an `exports` array with `params`, `columns`, and
    `business_object_type` (or legacy `object_type`).
- Python environment with `pandas` and `openpyxl` installed.

Output
- Updated JSON config file (same input path) with appended missing export attributes.
- Excel report (`.xlsx`) next to input CSV:
    - Yellow cells: attributes found in CSV but missing in JSON (added to JSON).
    - Red cells: attributes found in JSON but missing in CSV (added to Excel row).
    - Red `object_type` cell: no matching export object in JSON.
- Unmatched object-type report CSV (`*_unmatched_object_types.csv`).

Start Parameter
- `--csv` (optional): explicit path to `csv_column_headers_by_object_type_*.csv`.
- `--config` (optional): path to Vault loader JSON config.
- `--logs-dir` (optional): folder used to auto-detect latest aggregated CSV.

Function
- `find_latest_aggregated_csv`: resolves newest aggregated CSV when `--csv` is omitted.
- `build_export_index`: creates normalized lookup keys for export matching.
- `run_check`: performs comparison, synchronization, highlighting, and report generation.
- `write_excel_with_formatting`: writes formatted Excel review sheet.
- `write_unmatched_report`: writes object types missing in JSON mapping.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
RED_FILL = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")


def get_script_dir() -> Path:
    return Path(__file__).resolve().parent


def get_project_root() -> Path:
    script_dir = get_script_dir()
    for candidate in [script_dir, *script_dir.parents]:
        if (candidate / "config" / "vault_loader_config_basis_new.json").exists():
            return candidate
    raise FileNotFoundError("Could not locate project root containing config/vault_loader_config_basis_new.json")


def normalize_object_type(value: str) -> str:
    return re.sub(r"_+", "_", value.strip().lower())


def parse_export_object_from_params(params: str) -> str:
    match = re.search(r"(?:^|\s)-export\s+([^\s]+)", params or "")
    return match.group(1).strip() if match else ""


def find_latest_aggregated_csv(logs_dir: Path) -> Path:
    candidates = list(logs_dir.glob("csv_column_headers_by_object_type_*.csv"))
    if not candidates:
        raise FileNotFoundError(
            f"No files found for pattern: {logs_dir / 'csv_column_headers_by_object_type_*.csv'}"
        )
    return max(candidates, key=lambda item: item.stat().st_mtime)


def load_json_config(config_path: Path) -> dict:
    with config_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_json_config(config_path: Path, config: dict) -> None:
    with config_path.open("w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=4, ensure_ascii=False)


def get_head_columns(columns: list[str]) -> list[str]:
    head_columns = [column for column in columns if re.fullmatch(r"HEAD\d+", str(column))]
    return sorted(head_columns, key=lambda name: int(name[4:]))


def build_export_index(exports: list[dict]) -> dict[str, list[dict]]:
    index: dict[str, list[dict]] = {}
    for export in exports:
        keys: set[str] = set()
        object_type = str(export.get("business_object_type", export.get("object_type", ""))).strip()
        if object_type:
            keys.add(normalize_object_type(object_type))

        parsed_object = parse_export_object_from_params(str(export.get("params", "")))
        if parsed_object:
            keys.add(normalize_object_type(parsed_object))

        for key in keys:
            index.setdefault(key, []).append(export)
    return index


def ordered_union_columns(exports: list[dict]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for export in exports:
        for attribute in export.get("columns", []):
            text = str(attribute).strip()
            if text and text not in seen:
                seen.add(text)
                ordered.append(text)
    return ordered


def ensure_next_head_column(dataframe: pd.DataFrame, head_columns: list[str]) -> str:
    next_number = max((int(name[4:]) for name in head_columns), default=0) + 1
    next_column = f"HEAD{next_number}"
    dataframe[next_column] = ""
    head_columns.append(next_column)
    head_columns.sort(key=lambda name: int(name[4:]))
    return next_column


def write_excel_with_formatting(
    output_path: Path,
    dataframe: pd.DataFrame,
    yellow_cells: set[tuple[int, str]],
    red_cells: set[tuple[int, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "check_export_config"

    columns = list(dataframe.columns)
    worksheet.append(columns)

    for _, row in dataframe.iterrows():
        worksheet.append(["" if pd.isna(value) else value for value in row.tolist()])

    column_index = {name: position + 1 for position, name in enumerate(columns)}

    for row_index, column_name in yellow_cells:
        col = column_index.get(column_name)
        if col is not None:
            worksheet.cell(row=row_index + 2, column=col).fill = YELLOW_FILL

    for row_index, column_name in red_cells:
        col = column_index.get(column_name)
        if col is not None:
            worksheet.cell(row=row_index + 2, column=col).fill = RED_FILL

    workbook.save(output_path)


def write_unmatched_report(
    report_path: Path,
    unmatched_rows: list[dict[str, str]],
) -> None:
    report_dataframe = pd.DataFrame(unmatched_rows)
    if report_dataframe.empty:
        report_dataframe = pd.DataFrame(columns=["row_number", "object_type", "file_intent"])
    report_dataframe.to_csv(report_path, index=False, encoding="utf-8")


def run_check(csv_path: Path, config_path: Path) -> tuple[Path, Path]:
    dataframe = pd.read_csv(csv_path, dtype=str, encoding="utf-8").fillna("")
    config = load_json_config(config_path)

    exports = config.get("exports", [])
    if not isinstance(exports, list):
        raise ValueError("JSON config field 'exports' must be a list.")

    export_index = build_export_index(exports)
    head_columns = get_head_columns(list(dataframe.columns))

    yellow_cells: set[tuple[int, str]] = set()
    red_cells: set[tuple[int, str]] = set()

    unmatched_object_types = 0
    added_to_json = 0
    added_to_excel = 0
    unmatched_rows: list[dict[str, str]] = []

    for row_index in range(len(dataframe)):
        object_type_raw = str(dataframe.at[row_index, "object_type"]).strip()
        lookup_key = normalize_object_type(object_type_raw)

        matching_exports = export_index.get(lookup_key, [])
        if not matching_exports:
            red_cells.add((row_index, "object_type"))
            unmatched_object_types += 1
            unmatched_rows.append(
                {
                    "row_number": str(row_index + 2),
                    "object_type": object_type_raw,
                    "file_intent": str(dataframe.at[row_index, "file_intent"]).strip(),
                }
            )
            continue

        csv_attributes: list[str] = []
        attribute_cells: dict[str, list[str]] = {}
        for head in head_columns:
            value = str(dataframe.at[row_index, head]).strip()
            if not value:
                continue
            csv_attributes.append(value)
            attribute_cells.setdefault(value, []).append(head)

        json_attributes = ordered_union_columns(matching_exports)
        json_attribute_set = set(json_attributes)
        csv_attribute_set = set(csv_attributes)

        for attribute in csv_attributes:
            if attribute in json_attribute_set:
                continue
            for export in matching_exports:
                export_columns = export.setdefault("columns", [])
                if attribute not in export_columns:
                    export_columns.append(attribute)
                    added_to_json += 1
            for head in attribute_cells.get(attribute, []):
                yellow_cells.add((row_index, head))
            json_attribute_set.add(attribute)

        for attribute in json_attributes:
            if attribute in csv_attribute_set:
                continue

            if not head_columns:
                ensure_next_head_column(dataframe, head_columns)

            target_head = head_columns[-1]
            if str(dataframe.at[row_index, target_head]).strip():
                target_head = ensure_next_head_column(dataframe, head_columns)

            dataframe.at[row_index, target_head] = attribute
            red_cells.add((row_index, target_head))
            csv_attribute_set.add(attribute)
            added_to_excel += 1

    save_json_config(config_path, config)

    excel_path = csv_path.with_suffix(".xlsx")
    try:
        write_excel_with_formatting(excel_path, dataframe, yellow_cells, red_cells)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = csv_path.with_name(f"{csv_path.stem}_checked_{timestamp}.xlsx")
        write_excel_with_formatting(excel_path, dataframe, yellow_cells, red_cells)

    report_path = csv_path.with_name(f"{csv_path.stem}_unmatched_object_types.csv")
    write_unmatched_report(report_path, unmatched_rows)

    print(f"CSV source: {csv_path}")
    print(f"Excel output: {excel_path}")
    print(f"JSON updated: {config_path}")
    print(f"Unmatched object_type report: {report_path}")
    print(f"Rows: {len(dataframe)}")
    print(f"Object types not found in JSON (red): {unmatched_object_types}")
    print(f"Attributes added to JSON (yellow): {added_to_json}")
    print(f"Attributes added to Excel (red): {added_to_excel}")

    return excel_path, config_path


def parse_args() -> argparse.Namespace:
    project_root = get_project_root()
    default_logs_dir = project_root / "logs"
    default_config = project_root / "config" / "vault_loader_config_basis_new.json"

    parser = argparse.ArgumentParser()
    parser.add_argument("--csv", type=Path, default=None, help="Path to csv_column_headers_by_object_type_*.csv")
    parser.add_argument("--config", type=Path, default=default_config, help="Path to vault loader config JSON")
    parser.add_argument("--logs-dir", type=Path, default=default_logs_dir, help="Logs folder to search latest CSV")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    csv_path = args.csv if args.csv else find_latest_aggregated_csv(args.logs_dir)
    run_check(csv_path=csv_path, config_path=args.config)


if __name__ == "__main__":
    main()
